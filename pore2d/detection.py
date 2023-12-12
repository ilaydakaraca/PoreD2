#import torch
#import torchvision
#def detect_objects(image_path):
    # Load the YOLOv5 model
#    model = torch.hub.load('ultralytics/yolov5', 'custom', path='./weights/bestpore.pt')

#    # Load the image
#    image = torchvision.io.read_image(image_path)

    # Preprocess the image
#    image = image.unsqueeze(0).float() / 255.0  # Add batch dimension and normalize pixel values

    # Perform object detection
#    results = model(image)

    # Process the detection results as per your requirements
    # ...

#    return results

import os
from django.conf import settings
from django.shortcuts import render
from requests import request
import torch
import torchvision.transforms as transforms
from PIL import Image
import openpyxl
import pandas as pd
import os, shutil
import easyocr
import matplotlib.pyplot as plt
import cv2
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


# Load the YOLOv5 model
model = torch.hub.load('ultralytics/yolov5', 'custom', path='./weights/bestpore.pt')
model2 = torch.hub.load('ultralytics/yolov5', 'custom', path='./weights/best_scale_bar.pt')
model3 = torch.hub.load('ultralytics/yolov5', 'custom', path='./weights/best_window.pt')

# Define the object detection function
def detect_objects(image_filename):
    # Load the image
    absolute_path = os.path.join(settings.MEDIA_ROOT, image_filename)
    image = Image.open(absolute_path)
    
    # Convert the image to RGB if it has only one channel
    image = image.convert("RGB")

    output_folder = ''
    shutil.rmtree("pore")
    shutil.rmtree("media")
    results = model(image)
    #results.save(output_folder)
    results.save(output_folder, "pore") 
    results.print()
    results.pandas().xywh[0].to_excel("pore.xlsx")
    #deneme = results.show()
    reader = easyocr.Reader(['en']) # need to run only once to load model into memory
    result_text = reader.readtext('./pore/image0.jpg')
    #top_left = tuple(result_text[6][0][0])
    #bottom_right = tuple(result_text[6][0][2])
    text = result_text[7][1]

    with open('text.txt', 'w') as f:
        print(text, file=f)


    output_folder2 = ''
    shutil.rmtree("scale_bar")
    results2 = model2(image)
    results2.save(output_folder2, "scale_bar")
    results2.print()
    results2.pandas().xywh[0].to_excel("scalebar.xlsx")
    #deneme2 = results2.show()
    #context={'absolute_path':absolute_path,'results':results}
    #return render(request, 'index.html',context)
    #return render(request, 'detected_image.html', context)

    output_folder3 = '' 
    shutil.rmtree("window")
    results3 = model3(image)
    results3.save(output_folder3, "window")
    results3.print()
    results3.pandas().xywh[0].to_excel("window.xlsx")
    #deneme3 = results3.show()
    

    input_file = "pore.xlsx"
    input_file2 = "scalebar.xlsx"
    input_file3 = "window.xlsx"
    output_file = "output_pore.xlsx"
    output_file2 = "output_window.xlsx"
    text_file_path = "text.txt"

    with open(text_file_path, 'r') as text_file:
        text_data = text_file.read()

    input_wb = openpyxl.load_workbook(input_file)
    input_ws = input_wb.active

    input2_wb = openpyxl.load_workbook(input_file2)
    input2_ws = input2_wb.active

    input3_wb = openpyxl.load_workbook(input_file3)
    input3_ws = input3_wb.active

  
    
    input_ws['N1'] = 'Avg. Pore Size (px)'
    input_ws['O1'] = 'Scale Bar Size (px)'
    input_ws['P1'] = 'Avg. Pore Size (um)'
    input_ws['Q1'] = 'Avg. Pore Size (um) w/cf'
    input_ws['R1'] = 'Scale Bar Text'
    input_ws['S1'] = 'Scale Bar #'
    input_ws['T1'] = 'Scale Bar Unit'

    target_cell = input_ws['R2']  # Replace 'A1' with the desired cell reference
    target_cell.value = text_data

    total_rows = input_ws.max_row
    for row_number in range(total_rows, 1, -1):
        d_value = input_ws.cell(row=row_number, column=4).value  # D column (4th column) value
        e_value = input_ws.cell(row=row_number, column=5).value  # E column (5th column) value

        # Compare the values and get the larger one
        if d_value and e_value:
            larger_value = max(d_value, e_value)

            # Write the larger value to column K
            input_ws.cell(row=row_number, column=11, value=larger_value)

        # Delete the previous rows
        # Determine the total number of rows in the worksheet
    total_rows = input_ws.max_row

    # Initialize variables to store sum and count
    sum_k = 0
    count = 0

    # Iterate through rows in column K (excluding the header)
    for row_number in range(2, total_rows + 1):
        k_value = input_ws.cell(row=row_number, column=11).value  # K column (11th column) value
        if k_value is not None:
            sum_k += k_value
            count += 1

    # Calculate the average
    average_k = sum_k / count if count > 0 else 0

    # Write the average to a new cell (e.g., L2)
    input_ws['N2'] = average_k

    input_value = input2_ws['D2'].value
    input_ws['O2'] = input_value

    
    cell_value = input_ws['R2'].value
    if cell_value:
        part1 = cell_value[:-4]
        part2 = cell_value[-4:]

        input_ws['S2'].value = part1
        input_ws['T2'].value = part2
    
    # Extract values from cells L2 and M2
    n2_value = input_ws['N2'].value
    o2_value = input_ws['O2'].value
    s2_value = input_ws['S2'].value
    #p2_value = input_ws['P2'].value
    # Perform the desired calculations
    result = (n2_value / o2_value) * int(s2_value)
    input_ws['P2'] = result
    p2_value = input_ws['P2'].value
    result2 = (p2_value*2)/(1.7320508076)
    input_ws['Q2'] = result2

    original_column = input_ws['K']
    new_column1 = input_ws['L']
    new_column2 = input_ws['M']

    multiply_by1 = 2/1.7320508076
    multiply_by2 = (s2_value)
    divide_by = (o2_value)

    for cell in original_column:
        if cell.value is not None:  # Check for empty cells
            result3 = float(cell.value) * float(multiply_by1)
            new_column1[cell.row - 1].value = result3
            

    for cell in original_column:
        if cell.value is not None:  # Check for empty cells
            result4 = float(cell.value)/float(o2_value)*float(s2_value)* float(multiply_by1)
            new_column2[cell.row - 1].value = result4

    input_ws['K1'] = 'Pore Size (px)'
    input_ws['L1'] = 'Pore Size (w/Correction factor)'
    input_ws['M1'] = 'Pore Size (um, w/CF)'

    fill1 = PatternFill(start_color="FF5F1F", end_color="FF5F1F", fill_type="solid")
    fill_cell1 = input_ws['K1']
    fill_cell2 = input_ws['L1']
    fill_cell3 = input_ws['M1']
    fill_cell4 = input_ws['P1']
    fill_cell5 = input_ws['Q1']
    fill_cell1.fill = fill1
    fill_cell2.fill = fill1
    fill_cell3.fill = fill1
    fill_cell4.fill = fill1
    fill_cell5.fill = fill1
    fill2 = PatternFill(start_color="EBECF0", end_color="EBECF0", fill_type="solid")
    fill_cell6 = input_ws['N1']
    fill_cell7 = input_ws['O1']
    fill_cell8 = input_ws['R1']
    fill_cell9 = input_ws['S1']
    fill_cell10 = input_ws['T1']
    fill_cell6.fill = fill2
    fill_cell7.fill = fill2
    fill_cell8.fill = fill2
    fill_cell9.fill = fill2
    fill_cell10.fill = fill2

    for column in input_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column letter (e.g., 'A', 'B', 'C', ...)

        for cell in column:
            try:
                # Calculate the length of the cell's value
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        # Set the column width based on the maximum content length
        adjusted_width = (max_length)  # Add some padding
        input_ws.column_dimensions[column_letter].width = adjusted_width


    #for window
    
    input3_ws['M1'] = 'Avg. Window Size (px)'
    input3_ws['N1'] = 'Scale Bar Size (px)'
    input3_ws['O1'] = 'Avg. Window Size (um)'
    input3_ws['P1'] = 'Scale Bar Text'
    input3_ws['Q1'] = 'Scale Bar #'
    input3_ws['R1'] = 'Scale Bar Unit'

    target_cell = input3_ws['P2']  # Replace 'A1' with the desired cell reference
    target_cell.value = text_data

    total_rows = input3_ws.max_row
    for row_number in range(total_rows, 1, -1):
        d_value = input3_ws.cell(row=row_number, column=4).value  # D column (4th column) value
        e_value = input3_ws.cell(row=row_number, column=5).value  # E column (5th column) value

        # Compare the values and get the larger one
        if d_value and e_value:
            larger_value = max(d_value, e_value)

            # Write the larger value to column K
            input3_ws.cell(row=row_number, column=11, value=larger_value)

        # Delete the previous rows
        # Determine the total number of rows in the worksheet
    total_rows = input3_ws.max_row

    # Initialize variables to store sum and count
    sum_k = 0
    count = 0

    # Iterate through rows in column K (excluding the header)
    for row_number in range(2, total_rows + 1):
        k_value = input3_ws.cell(row=row_number, column=11).value  # K column (11th column) value
        if k_value is not None:
            sum_k += k_value
            count += 1

    # Calculate the average
    average_k = sum_k / count if count > 0 else 0

    # Write the average to a new cell (e.g., L2)
    input3_ws['M2'] = average_k

    input_value = input2_ws['D2'].value
    input3_ws['N2'] = input_value

    
    cell_value = input3_ws['P2'].value
    if cell_value:
        part1 = cell_value[:-4]
        part2 = cell_value[-4:]

        input3_ws['Q2'].value = part1
        input3_ws['R2'].value = part2
    
    # Extract values from cells L2 and M2
    m2_value = input3_ws['M2'].value
    n2_value = input3_ws['N2'].value
    q2_value = input3_ws['Q2'].value
    #p2_value = input_ws['P2'].value
    # Perform the desired calculations
    result = (m2_value / n2_value) * int(q2_value)
    input3_ws['O2'] = result


    original_column = input3_ws['K']
    new_column1 = input3_ws['L']

    for cell in original_column:
        if cell.value is not None:  # Check for empty cells
            result3 = float(cell.value)/float(n2_value) * float(q2_value)
            new_column1[cell.row - 1].value = result3

    input3_ws['K1'] = 'Window Size (px)'
    input3_ws['L1'] = 'Window_Size (um)'

    fill3 = PatternFill(start_color="FF5F1F", end_color="FF5F1F", fill_type="solid")
    fill_cell11 = input3_ws['K1']
    fill_cell12 = input3_ws['L1']
    fill_cell13 = input3_ws['O1']
    fill_cell11.fill = fill3
    fill_cell12.fill = fill3
    fill_cell13.fill = fill3

    fill4 = PatternFill(start_color="EBECF0", end_color="EBECF0", fill_type="solid")
    fill_cell14 = input3_ws['M1']
    fill_cell15 = input3_ws['N1']
    fill_cell16 = input3_ws['P1']
    fill_cell17 = input3_ws['Q1']
    fill_cell18 = input3_ws['R1']
    fill_cell14.fill = fill4
    fill_cell15.fill = fill4
    fill_cell16.fill = fill4
    fill_cell17.fill = fill4
    fill_cell18.fill = fill4
    

    for column in input3_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column letter (e.g., 'A', 'B', 'C', ...)

        for cell in column:
            try:
                # Calculate the length of the cell's value
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        # Set the column width based on the maximum content length
        adjusted_width = (max_length)  # Add some padding
        input3_ws.column_dimensions[column_letter].width = adjusted_width


    input_wb.save(output_file)
    input3_wb.save(output_file2)

    return results, input_wb,  text
